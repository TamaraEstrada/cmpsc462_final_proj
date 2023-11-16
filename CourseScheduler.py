from openpyxl import load_workbook


class CourseScheduler:
    def __init__(self, available_classes, schedule_file):
        self.available = available_classes
        self.taken = self.load_taken_classes(schedule_file)
        self.current = {}

    def load_taken_classes(self, schedule_file):
        workbook = load_workbook(filename=schedule_file)
        sheet = workbook.active
        taken = {}
        for row in sheet.iter_rows(values_only=True):
            course = row[0]
            course_info = {"credits": row[1], "credit_type": row[2]}
            taken[course] = course_info
        return taken

    def schedule_course(self, sheet):
        course = input("What course would you like to schedule? ")
        print(course)
        if course in self.available:
            if course not in self.taken:
                semester = input("Fall or Spring semester? ")
                if semester in self.available[course]["semesters"]:
                    current_credits = self.available[course]["credits"]
                    for c in self.current:
                        current_credits += self.available[c]["credits"]
                    print(current_credits)
                    if current_credits < 19:
                        valid = True
                        for prereq in self.available[course]["prerequisites"]:
                            if prereq not in self.taken and prereq is not None:
                                valid = False
                                print(f"{prereq} has not been taken/scheduled yet.")
                        if valid:
                            credits = self.available[course]["credits"]
                            type = self.available[course]["credit_type"]
                            print(course, credits, type, sep="\t")
                            sheet.insert_rows(idx=1)
                            sheet["A1"] = course
                            sheet["B1"] = credits
                            sheet["C1"] = type
                            workbook.save("TestSchedule.xlsx")
                            self.taken[course] = self.available[course]
                            self.current[course] = self.available[course]
                            print("Course scheduled!")
                    else:
                        print("Course would exceed credit limit for semester.")
                else:
                    print("Course unavailable in given semester.")
            else:
                print("Course has already been scheduled/taken.")
        else:
            print("Invalid course name.")

    def run_scheduler(self, sheet):
        while True:
            x = input("Would you like to schedule a class? ")
            if x.lower() == "yes":
                self.schedule_course(sheet)
            else:
                print("Okay.")
                break
