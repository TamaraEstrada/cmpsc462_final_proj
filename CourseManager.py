from openpyxl import load_workbook


class CourseManager:
    @staticmethod
    def load_courses(filename):
        workbook = load_workbook(filename=filename)
        sheet = workbook.active
        courses = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            course = row[0]
            courses[course] = {
                "credits": row[1],
                "credit_type": row[2],
                "semesters": list(row[3].split(", ")),
                # list comprehension to include only non-None values
                "prerequisites": [prereq for prereq in row[4:] if prereq is not None],
            }
        # print(available_classes)
        for course, info in courses.items():
            print(course, info, sep=": ")
            print()  # This will print a new line after each entry

        return courses
